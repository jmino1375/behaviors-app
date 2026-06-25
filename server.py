import io
import json
import os
import re
import traceback
from datetime import datetime

import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import FastAPI, WebSocket, WebSocketDisconnect
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

app = FastAPI()

# Startup check
api_key = os.environ.get("ANTHROPIC_API_KEY", "")
if api_key:
    print(f"ANTHROPIC_API_KEY found: sk-ant-...{api_key[-4:]}")
else:
    print("WARNING: ANTHROPIC_API_KEY is NOT set!")

# ── Persistence ────────────────────────────────────────────────────────────
DATA_FILE = os.environ.get("DATA_FILE", "data.json")

def load_data():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return {"behaviors": {"above": [], "below": []}, "categories": {"above": [], "below": []}}

def save_data():
    with open(DATA_FILE, "w") as f:
        json.dump({"behaviors": behaviors, "categories": categories}, f, ensure_ascii=False, indent=2)

_loaded = load_data()
behaviors: dict[str, list[str]] = _loaded["behaviors"]
categories: dict[str, list] = _loaded["categories"]
clients: list[WebSocket] = []


# ── WebSocket broadcast ────────────────────────────────────────────────────
async def broadcast(msg: dict):
    dead = []
    for ws in clients:
        try:
            await ws.send_text(json.dumps(msg))
        except Exception:
            dead.append(ws)
    for ws in dead:
        clients.remove(ws)


# ── WebSocket endpoint ─────────────────────────────────────────────────────
@app.websocket("/ws")
async def websocket_endpoint(ws: WebSocket):
    await ws.accept()
    clients.append(ws)
    await ws.send_text(json.dumps({
        "type": "state",
        "behaviors": behaviors,
        "categories": categories,
    }))
    try:
        while True:
            data = await ws.receive_text()
            msg = json.loads(data)
            if msg.get("type") == "submit":
                kind = msg.get("kind")
                text = (msg.get("text") or "").strip()[:500]
                if kind in ("above", "below") and text:
                    behaviors[kind].append(text)
                    save_data()
                    await broadcast({"type": "new_behavior", "kind": kind, "text": text})
    except WebSocketDisconnect:
        if ws in clients:
            clients.remove(ws)


# ── Categorize endpoint ────────────────────────────────────────────────────
@app.post("/categorize")
async def categorize_behaviors():
    try:
        above_list = "\n".join(f"{i+1}. {b}" for i, b in enumerate(behaviors["above"])) or "(ninguno)"
        below_list = "\n".join(f"{i+1}. {b}" for i, b in enumerate(behaviors["below"])) or "(ninguno)"

        prompt = f"""Analiza los siguientes comportamientos y agrúpalos en categorías temáticas comunes. Responde ÚNICAMENTE con JSON válido.

COMPORTAMIENTOS SOBRE LA LÍNEA:
{above_list}

COMPORTAMIENTOS BAJO LA LÍNEA:
{below_list}

Responde con este formato JSON exacto:
{{
  "above": [
    {{ "category": "Nombre de categoría", "items": ["comportamiento 1", "comportamiento 2"] }}
  ],
  "below": [
    {{ "category": "Nombre de categoría", "items": ["comportamiento 1", "comportamiento 2"] }}
  ]
}}

Reglas:
- Agrupa comportamientos similares o relacionados bajo una misma categoría
- Los nombres de categoría deben ser descriptivos y en español
- Cada comportamiento debe aparecer en exactamente una categoría
- Si no hay comportamientos en un tipo, devuelve array vacío"""

        ai_client = anthropic.Anthropic()
        message = ai_client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}],
        )
        text = message.content[0].text.strip()
        match = re.search(r"\{[\s\S]*\}", text)
        if not match:
            return JSONResponse(status_code=500, content={"error": "No JSON in Claude response", "raw": text[:200]})

        global categories
        categories = json.loads(match.group())
        save_data()
        await broadcast({"type": "categories_updated", "categories": categories})
        return {"success": True, "categories": categories}
    except Exception as e:
        tb = traceback.format_exc()
        print(f"CATEGORIZE ERROR: {tb}")
        return JSONResponse(status_code=500, content={"error": str(e), "detail": tb[-500:]})


# ── Download endpoint ──────────────────────────────────────────────────────
@app.get("/download")
def download_data():
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    wb = openpyxl.Workbook()

    green_fill = PatternFill("solid", fgColor="1E4620")
    red_fill   = PatternFill("solid", fgColor="4C1515")
    green_hdr  = PatternFill("solid", fgColor="2D6A31")
    red_hdr    = PatternFill("solid", fgColor="7B2020")
    bold       = Font(bold=True, color="FFFFFF")
    white      = Font(color="FFFFFF")
    wrap       = Alignment(wrap_text=True, vertical="top")

    def make_sheet(ws, kind, fill, hdr_fill):
        ws.column_dimensions["A"].width = 60
        ws.column_dimensions["B"].width = 30
        # Header
        ws["A1"] = "Comportamiento"
        ws["B1"] = "Categoría"
        for cell in [ws["A1"], ws["B1"]]:
            cell.font = bold
            cell.fill = hdr_fill
            cell.alignment = wrap

        # Build category lookup
        cat_map = {}
        for cat in (categories.get(kind) or []):
            for item in cat.get("items", []):
                cat_map[item] = cat.get("category", "")

        items = behaviors[kind]
        for i, text in enumerate(items, start=2):
            ws.cell(i, 1, text).fill = fill
            ws.cell(i, 1).font = white
            ws.cell(i, 1).alignment = wrap
            ws.cell(i, 2, cat_map.get(text, "")).fill = fill
            ws.cell(i, 2).font = white
            ws.cell(i, 2).alignment = wrap

    # Sheet 1: Sobre la línea
    ws1 = wb.active
    ws1.title = "Sobre la línea"
    make_sheet(ws1, "above", green_fill, green_hdr)

    # Sheet 2: Bajo la línea
    ws2 = wb.create_sheet("Bajo la línea")
    make_sheet(ws2, "below", red_fill, red_hdr)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"comportamientos_{now}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── Clear endpoint ─────────────────────────────────────────────────────────
@app.post("/clear")
async def clear_data():
    global categories
    behaviors["above"].clear()
    behaviors["below"].clear()
    categories = {"above": [], "below": []}
    save_data()
    await broadcast({"type": "cleared"})
    return {"success": True}


# ── Static files ───────────────────────────────────────────────────────────
app.mount("/", StaticFiles(directory="public", html=True), name="static")


# ── Run ────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 3000))
    uvicorn.run(app, host="0.0.0.0", port=port)
